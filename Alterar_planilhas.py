import openpyxl
import random
import os

# Arquivos para processar
arquivos = {
    'GLPI - Chamados.xlsx': 'GLPI - Chamados ALTERADA.xlsx',
    'GLPI - Chamados(SLA NORMAL).xlsx': 'GLPI - Chamados(SLA NORMAL)ALTERADA.xlsx',
    'GLPI - Chamados  (SLA PENDENTE).xlsx': 'GLPI - Chamados(SLA PENDENTE)ALTERADA.xlsx',
    'GLPI - Chamados(SLA URGENTE).xlsx': 'GLPI - Chamados(SLA URGENTE)ALTERADA.xlsx'
}

# Dados para substituição
tecnicos = ['TEC Joana Silva', 'TEC Jonas Silveira', 'TEC Pedro Silva', 'TEC João Santos', 'TEC Maria Silva']
empresas_ficticias = [
    'EMPRESA X', 'EMPRESA Y', 'EMPRESA Z', 'EMPRESA C',
    'EMPRESA B', 'EMPRESA A', 'EMPRESA G', 'EMPRESA F', 'EMPRESA E'
]

# Processa cada arquivo da lista informada
for entrada, saida in arquivos.items():
    if not os.path.exists(entrada):
        print(f"Arquivo não encontrado: {entrada}")
        continue

    print(f"Processando: {entrada}")
    workbook = openpyxl.load_workbook(entrada)
    sheet = workbook.active

    # TRATATIVA DIFERENTE PARA ESTA PLANILHA POIS AS COLUNAS SÃO DIFERENTES
    if (entrada == 'GLPI - Chamados.xlsx'):
        sheet['J1'].value = 'Técnico'
        for row in range(2, sheet.max_row + 1):
            sheet[f'J{row}'].value = random.choice(tecnicos)
        
    if (entrada == 'GLPI - Chamados.xlsx'):
        for row in range(2, sheet.max_row + 1):
            celula = sheet[f'I{row}']
            if celula.value:
                celula.value = random.choice(empresas_ficticias)

    # Substitui nomes na coluna F, ignorando a primeira linha (header)
    if entrada in (
        'GLPI - Chamados(SLA NORMAL).xlsx',
        'GLPI - Chamados  (SLA PENDENTE).xlsx',
        'GLPI - Chamados(SLA URGENTE).xlsx'
    ):
        sheet['F1'].value = 'Técnico'
        for row in range(2, sheet.max_row + 1):
            sheet[f'F{row}'].value = random.choice(tecnicos)
        
    # Substitui empresas na coluna E, ignorando a primeira linha (header)
    if entrada in (
        'GLPI - Chamados(SLA NORMAL).xlsx',
        'GLPI - Chamados  (SLA PENDENTE).xlsx',
        'GLPI - Chamados(SLA URGENTE).xlsx'
    ):
        for row in range(2, sheet.max_row + 1):
            celula = sheet[f'E{row}']
            if celula.value:
                celula.value = random.choice(empresas_ficticias)

    workbook.save(saida)
    print(f"Arquivo salvo como: {saida}")

print("Processamento concluído.")
